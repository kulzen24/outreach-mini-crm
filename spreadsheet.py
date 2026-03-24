"""
spreadsheet.py
All reads and writes go through outreach.xlsx, which doubles as the live checklist.
Open it anytime in Excel or Numbers to see current status.

Column layout (A → K):
  A  station_name
  B  contact_name
  C  email
  D  school
  E  city
  F  state
  G  genre
  H  notes
  I  status        (pending / sent / skipped)
  J  sent_at
  K  row_id        (hidden stable identifier)
"""

import os
import uuid
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX_PATH = os.path.join(os.path.dirname(__file__), "outreach.xlsx")

COLUMNS = [
    "station_name",
    "contact_name",
    "email",
    "school",
    "city",
    "state",
    "genre",
    "notes",
    "status",
    "sent_at",
    "row_id",
]

HEADERS = [
    "Station Name",
    "Contact / Music Director",
    "Email",
    "School",
    "City",
    "State",
    "Genre / Format",
    "Notes",
    "Status",
    "Sent At",
    "ID",
]

# Status → fill colour
STATUS_FILLS = {
    "sent":    PatternFill("solid", fgColor="C6EFCE"),   # green
    "skipped": PatternFill("solid", fgColor="FFEB9C"),   # yellow
    "pending": PatternFill("solid", fgColor="FFFFFF"),   # white
}

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COL_WIDTHS = [30, 25, 35, 30, 18, 8, 20, 30, 10, 20, 0]  # 0 = hidden


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _load_or_create_wb():
    if os.path.exists(XLSX_PATH):
        return openpyxl.load_workbook(XLSX_PATH)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outreach"
    _write_header_row(ws)
    wb.save(XLSX_PATH)
    return wb


def _write_header_row(ws):
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        col_letter = get_column_letter(col_idx)
        if width == 0:
            ws.column_dimensions[col_letter].hidden = True
        else:
            ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _ws(wb):
    return wb["Outreach"] if "Outreach" in wb.sheetnames else wb.active


def _row_to_dict(ws, row_idx):
    d = {}
    for col_idx, key in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        d[key] = cell.value if cell.value is not None else ""
    return d


def _style_data_row(ws, row_idx, status="pending"):
    fill = STATUS_FILLS.get(status, STATUS_FILLS["pending"])
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)


def _find_row_by_id(ws, row_id):
    id_col = COLUMNS.index("row_id") + 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=id_col).value == row_id:
            return row_idx
    return None


def _save(wb):
    wb.save(XLSX_PATH)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def init_spreadsheet():
    """Create outreach.xlsx with headers if it doesn't exist."""
    _load_or_create_wb()


def upsert_stations(records):
    """
    Add records to the spreadsheet, skipping rows where
    (station_name + email) already exist. Returns count inserted.
    """
    wb = _load_or_create_wb()
    ws = _ws(wb)

    # Build a set of existing (station_name, email) pairs for dedup
    existing = set()
    for row_idx in range(2, ws.max_row + 1):
        name = str(ws.cell(row=row_idx, column=1).value or "").strip().lower()
        email = str(ws.cell(row=row_idx, column=3).value or "").strip().lower()
        if name or email:
            existing.add((name, email))

    inserted = 0
    for r in records:
        key = (
            str(r.get("station_name", "")).strip().lower(),
            str(r.get("email", "")).strip().lower(),
        )
        if key in existing:
            continue

        row_idx = ws.max_row + 1
        values = [
            r.get("station_name", ""),
            r.get("contact_name", ""),
            r.get("email", ""),
            r.get("school", ""),
            r.get("city", ""),
            r.get("state", ""),
            r.get("genre", ""),
            r.get("notes", ""),
            "pending",
            "",
            str(uuid.uuid4()),
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        _style_data_row(ws, row_idx, "pending")
        existing.add(key)
        inserted += 1

    _save(wb)
    return inserted


def get_stations(status=None, search=None):
    """Return list of station dicts, optionally filtered."""
    wb = _load_or_create_wb()
    ws = _ws(wb)

    results = []
    for row_idx in range(2, ws.max_row + 1):
        d = _row_to_dict(ws, row_idx)
        if not any(d.values()):
            continue
        if status and status != "all" and d.get("status") != status:
            continue
        if search:
            term = search.lower()
            if not any(term in str(v).lower() for v in d.values()):
                continue
        results.append(d)

    return results


def get_station(row_id):
    """Return a single station dict by its row_id."""
    wb = _load_or_create_wb()
    ws = _ws(wb)
    row_idx = _find_row_by_id(ws, row_id)
    return _row_to_dict(ws, row_idx) if row_idx else None


def update_station(row_id, fields):
    """Update specific fields on a station row."""
    wb = _load_or_create_wb()
    ws = _ws(wb)
    row_idx = _find_row_by_id(ws, row_id)
    if row_idx is None:
        return

    allowed = {k for k in COLUMNS if k != "row_id"}
    for key, val in fields.items():
        if key not in allowed:
            continue
        col_idx = COLUMNS.index(key) + 1
        ws.cell(row=row_idx, column=col_idx, value=val)

    new_status = fields.get("status") or str(ws.cell(row=row_idx, column=COLUMNS.index("status") + 1).value or "pending")
    _style_data_row(ws, row_idx, new_status)
    _save(wb)


def mark_sent(row_id):
    """Set status → sent and record the timestamp."""
    update_station(row_id, {
        "status": "sent",
        "sent_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    })


def mark_skipped(row_id):
    update_station(row_id, {"status": "skipped"})


def mark_pending(row_id):
    update_station(row_id, {"status": "pending", "sent_at": ""})


def delete_station(row_id):
    """Remove a row from the spreadsheet entirely."""
    wb = _load_or_create_wb()
    ws = _ws(wb)
    row_idx = _find_row_by_id(ws, row_id)
    if row_idx:
        ws.delete_rows(row_idx)
        _save(wb)


def get_stats():
    rows = get_stations()
    total = len(rows)
    sent = sum(1 for r in rows if r.get("status") == "sent")
    skipped = sum(1 for r in rows if r.get("status") == "skipped")
    pending = sum(1 for r in rows if r.get("status") == "pending")
    return {"total": total, "sent": sent, "skipped": skipped, "pending": pending}
